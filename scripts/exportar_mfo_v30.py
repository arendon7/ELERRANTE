#!/usr/bin/env python3
"""Exporta el MFO v3.3 de El Errante a un snapshot JSON V3.0.

El workbook y el snapshot permanecen locales. El script valida el perfil exacto
observado en MFO_EL_ERRANTE_v3_3_Decisiones_y_Escenarios y falla si la
estructura cambia. No infiere columnas por posición sin comprobar etiquetas.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

SCHEMA_VERSION = "3.0"
WORKBOOK_PROFILE = "MFO_V3_3_DECISIONES_ESCENARIOS"
REQUIRED_SHEETS = (
    "00_INICIO",
    "05_PRODUCTOS_SUPUESTOS",
    "01_PLAN_VENTAS",
    "02_PRODUCCION_COMPRAS",
    "03_RESULTADOS_CAJA",
    "04_DASHBOARD",
    "06_AUDITORIA",
    "07_REAL_VS_PLAN",
    "08_DECISIONES_ESCENARIOS",
)
VALID_STATES = {"CONFIRMADO", "ESTIMADO", "INFERIDO", "CONTRADICTORIO", "PENDIENTE"}


def die(message: str) -> None:
    raise SystemExit(message)


def load_openpyxl():
    try:
        import openpyxl  # type: ignore
        from openpyxl.utils.datetime import from_excel  # type: ignore
    except ImportError:
        die("Falta openpyxl. Instala localmente con: python3 -m pip install openpyxl")
    return openpyxl, from_excel


def clean(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_quality(raw: Any) -> str:
    value = as_text(raw).upper()
    if value in {"APROBADO", "CONFIRMADO"} or value.startswith("OFICIAL"):
        return "CONFIRMADO"
    if value in {"CALCULADO", "INFERIDO", "CONFIRMADO MODELO"}:
        return "INFERIDO"
    if value in {"PROVISIONAL", "CONFIRMADO PARCIAL"}:
        return "ESTIMADO"
    if value in {"EDITABLE", "DECISIÓN", "PENDIENTE"}:
        return "PENDIENTE"
    if value in {"INTENCIONAL", "PASA"}:
        return "CONFIRMADO"
    return "ESTIMADO"


def confidence_for(status: str) -> str:
    return "Alta" if status == "CONFIRMADO" else "Media"


def month_value(value: Any, from_excel) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).strftime("%Y-%m")
        except Exception:
            pass
    text = as_text(value)
    if len(text) >= 7 and text[4] in "-/":
        return text[:7].replace("/", "-")
    die(f"No se pudo interpretar periodo mensual: {value!r}")


def checked_value(ws_values, ws_formulas, row: int, col: int) -> Any:
    value = ws_values.cell(row=row, column=col).value
    formula = ws_formulas.cell(row=row, column=col).value
    if isinstance(formula, str) and formula.startswith("=") and value is None:
        coord = ws_values.cell(row=row, column=col).coordinate
        die(
            f"{ws_values.title}!{coord}: fórmula sin valor calculado almacenado. "
            "Abre y guarda el workbook en Excel o LibreOffice antes de exportar."
        )
    return clean(value)


def require_cell(ws, coordinate: str, expected: str) -> None:
    found = as_text(ws[coordinate].value)
    if found != expected:
        die(f"{ws.title}!{coordinate}: se esperaba {expected!r} y se encontró {found!r}.")


def header_map(ws, row: int, start_col: int, end_col: int) -> dict[str, int]:
    result: dict[str, int] = {}
    duplicates: set[str] = set()
    for col in range(start_col, end_col + 1):
        label = as_text(ws.cell(row=row, column=col).value)
        if not label:
            continue
        if label in result:
            duplicates.add(label)
        result[label] = col
    if duplicates:
        die(f"{ws.title} fila {row}: encabezados duplicados: {', '.join(sorted(duplicates))}")
    return result


def require_headers(headers: dict[str, int], required: tuple[str, ...], where: str) -> None:
    missing = [label for label in required if label not in headers]
    if missing:
        die(f"{where}: faltan encabezados: {', '.join(missing)}")


def validate_profile(wb) -> None:
    missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
    if missing:
        die("El archivo no corresponde al perfil MFO v3.3. Faltan hojas: " + ", ".join(missing))

    require_cell(wb["00_INICIO"], "A1", "EL ERRANTE — MFO v3 · Planeación y Caja")
    require_cell(wb["05_PRODUCTOS_SUPUESTOS"], "A5", "SKU")
    require_cell(wb["05_PRODUCTOS_SUPUESTOS"], "B5", "Producto")
    require_cell(wb["01_PLAN_VENTAS"], "A5", "SKU")
    require_cell(wb["01_PLAN_VENTAS"], "A35", "SKU")
    require_cell(wb["03_RESULTADOS_CAJA"], "A25", "Flujo de caja y disponibilidad")
    require_cell(wb["06_AUDITORIA"], "A22", "Hallazgos y decisiones pendientes")
    require_cell(wb["08_DECISIONES_ESCENARIOS"], "A5", "Decisión")
    require_cell(wb["08_DECISIONES_ESCENARIOS"], "A35", "Escenarios de sensibilidad — año 1")
    require_cell(wb["08_DECISIONES_ESCENARIOS"], "A36", "Escenario")


def inspect_profile(wb) -> dict[str, Any]:
    return {
        "profile": WORKBOOK_PROFILE,
        "sheets": list(wb.sheetnames),
        "anchors": {
            "products": "05_PRODUCTOS_SUPUESTOS!A5:W19",
            "planYear1": "01_PLAN_VENTAS!A5:P19",
            "planYear2": "01_PLAN_VENTAS!A35:P49",
            "cashFlow": "03_RESULTADOS_CAJA!A25:Y35",
            "auditPending": "06_AUDITORIA!A22:J29",
            "decisions": "08_DECISIONES_ESCENARIOS!A5:H10",
            "scenarios": "08_DECISIONES_ESCENARIOS!A35:K40",
        },
    }


def product_master(wbv, wbf) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    v = wbv["05_PRODUCTOS_SUPUESTOS"]
    f = wbf["05_PRODUCTOS_SUPUESTOS"]
    headers = header_map(f, 5, 1, 23)
    required = (
        "SKU", "Producto", "Categoría", "Precio final (COP)", "Costo directo (COP)",
        "Estado", "Fuente"
    )
    require_headers(headers, required, "05_PRODUCTOS_SUPUESTOS fila 5")
    rows: list[dict[str, Any]] = []
    by_sku: dict[str, dict[str, Any]] = {}
    row = 6
    while row <= v.max_row:
        sku = as_text(checked_value(v, f, row, headers["SKU"]))
        if not sku:
            break
        model_status = as_text(checked_value(v, f, row, headers["Estado"]))
        status = normalize_quality(model_status)
        item = {
            "sku": sku,
            "name": checked_value(v, f, row, headers["Producto"]),
            "category": checked_value(v, f, row, headers["Categoría"]),
            "price": checked_value(v, f, row, headers["Precio final (COP)"]),
            "directCost": checked_value(v, f, row, headers["Costo directo (COP)"]),
            "validFrom": "2026-09-01",
            "status": status,
            "modelStatus": model_status,
            "confidence": confidence_for(status),
            "source": checked_value(v, f, row, headers["Fuente"]) or "05_PRODUCTOS_SUPUESTOS",
        }
        if item["price"] in (None, "") or item["directCost"] in (None, ""):
            die(f"05_PRODUCTOS_SUPUESTOS fila {row}: precio y costo directo son obligatorios.")
        rows.append(item)
        by_sku[sku] = item
        row += 1
    if len(rows) != 14:
        die(f"Perfil v3.3: se esperaban 14 SKU en el maestro y se encontraron {len(rows)}.")
    return rows, by_sku


def plan_block(wbv, wbf, header_row: int, products: dict[str, dict[str, Any]], from_excel) -> list[dict[str, Any]]:
    v = wbv["01_PLAN_VENTAS"]
    f = wbf["01_PLAN_VENTAS"]
    if as_text(f.cell(header_row, 1).value) != "SKU" or as_text(f.cell(header_row, 2).value) != "Producto":
        die(f"01_PLAN_VENTAS fila {header_row}: bloque de ventas no reconocido.")
    months = [month_value(checked_value(v, f, header_row, col), from_excel) for col in range(3, 15)]
    result: list[dict[str, Any]] = []
    row = header_row + 1
    seen = 0
    while row <= v.max_row:
        sku = as_text(checked_value(v, f, row, 1))
        if not sku:
            break
        if sku not in products:
            die(f"01_PLAN_VENTAS fila {row}: SKU {sku!r} no existe en el maestro.")
        product = products[sku]
        for offset, col in enumerate(range(3, 15)):
            qty = checked_value(v, f, row, col)
            if qty in (None, ""):
                die(f"01_PLAN_VENTAS fila {row}, periodo {months[offset]}: cantidad vacía.")
            price = product["price"]
            cost = product["directCost"]
            result.append({
                "month": months[offset],
                "sku": sku,
                "quantity": qty,
                "unitPrice": price,
                "sales": qty * price,
                "unitCost": cost,
                "cogs": qty * cost,
                "status": "ESTIMADO",
                "confidence": "Media",
                "source": "01_PLAN_VENTAS",
            })
        seen += 1
        row += 1
    if seen != 14:
        die(f"01_PLAN_VENTAS fila {header_row}: se esperaban 14 SKU y se encontraron {seen}.")
    return result


def cash_flow(wbv, wbf, from_excel) -> list[dict[str, Any]]:
    v = wbv["03_RESULTADOS_CAJA"]
    f = wbf["03_RESULTADOS_CAJA"]
    months = [month_value(checked_value(v, f, 4, col), from_excel) for col in range(2, 26)]
    labels = {
        "Caja inicial": "openingCash",
        "Ventas cobradas": "salesCash",
        "Compras pagadas": "purchases",
        "Gastos operativos pagados": "operatingExpenses",
        "Auxiliares pagados": "auxiliaryPayroll",
        "Juan pagado": "juanCash",
        "Impuestos reservados": "taxReserve",
        "Arriendo pagado": "rent",
        "CAPEX": "capex",
        "Caja final": "endingCash",
    }
    rows_by_label: dict[str, int] = {}
    for row in range(26, 36):
        label = as_text(f.cell(row, 1).value)
        if label:
            rows_by_label[label] = row
    missing = [label for label in labels if label not in rows_by_label]
    if missing:
        die("03_RESULTADOS_CAJA: faltan filas de caja: " + ", ".join(missing))

    result: list[dict[str, Any]] = []
    for idx, col in enumerate(range(2, 26)):
        item: dict[str, Any] = {
            "month": months[idx],
            "status": "ESTIMADO",
            "confidence": "Media",
            "source": "03_RESULTADOS_CAJA",
        }
        for label, field in labels.items():
            item[field] = checked_value(v, f, rows_by_label[label], col)
        result.append(item)
    return result


def add_assumption(
    output: list[dict[str, Any]],
    name: Any,
    value: Any,
    unit: Any,
    model_status: Any,
    note: Any,
    category: str,
    **extra: Any,
) -> None:
    if not as_text(name):
        return
    status = normalize_quality(model_status)
    item = {
        "name": clean(name),
        "value": clean(value),
        "unit": clean(unit),
        "status": status,
        "modelStatus": clean(model_status),
        "confidence": confidence_for(status),
        "source": "05_PRODUCTOS_SUPUESTOS",
        "note": clean(note),
        "category": category,
    }
    item.update({k: clean(v) for k, v in extra.items()})
    output.append(item)


def assumptions(wbv, wbf) -> list[dict[str, Any]]:
    v = wbv["05_PRODUCTOS_SUPUESTOS"]
    f = wbf["05_PRODUCTOS_SUPUESTOS"]
    output: list[dict[str, Any]] = []

    for row in range(24, 33):
        add_assumption(
            output,
            checked_value(v, f, row, 1), checked_value(v, f, row, 2),
            checked_value(v, f, row, 3), checked_value(v, f, row, 4),
            checked_value(v, f, row, 5), "Caja, impuestos y compras"
        )
        add_assumption(
            output,
            checked_value(v, f, row, 7), checked_value(v, f, row, 8),
            checked_value(v, f, row, 9), checked_value(v, f, row, 10),
            checked_value(v, f, row, 11), "Personal y gastos"
        )
        decision = checked_value(v, f, row, 13)
        if decision not in (None, ""):
            add_assumption(
                output,
                decision, checked_value(v, f, row, 15), "costo / factor",
                checked_value(v, f, row, 16), checked_value(v, f, row, 17),
                "Decisiones de crecimiento",
                configuredMonth=checked_value(v, f, row, 14),
            )

    for row in (33, 34):
        add_assumption(
            output,
            checked_value(v, f, row, 7), checked_value(v, f, row, 8),
            checked_value(v, f, row, 9), checked_value(v, f, row, 10),
            checked_value(v, f, row, 11), "Personal y gastos"
        )

    for row in range(39, 43):
        stage = checked_value(v, f, row, 1)
        if stage not in (None, ""):
            add_assumption(
                output,
                stage, checked_value(v, f, row, 3), "unidades físicas / mes",
                checked_value(v, f, row, 4), checked_value(v, f, row, 5),
                "Capacidad productiva",
                fromMonth=checked_value(v, f, row, 2),
            )
        add_assumption(
            output,
            checked_value(v, f, row, 7), checked_value(v, f, row, 8),
            checked_value(v, f, row, 9), checked_value(v, f, row, 10),
            checked_value(v, f, row, 11), "Parámetros de producción"
        )

    for row in range(46, 50):
        component = checked_value(v, f, row, 7)
        if component not in (None, ""):
            reference = checked_value(v, f, row, 9)
            add_assumption(
                output,
                component, checked_value(v, f, row, 8),
                checked_value(v, f, row, 10), checked_value(v, f, row, 11),
                f"Costo de referencia: {reference}", "Costos sensibles"
            )
        add_assumption(
            output,
            checked_value(v, f, row, 13), checked_value(v, f, row, 14),
            checked_value(v, f, row, 15), checked_value(v, f, row, 16),
            checked_value(v, f, row, 17), "Política de pago de Juan"
        )

    if len(output) != 44:
        die(f"Perfil v3.3: se esperaban 44 supuestos normalizados y se encontraron {len(output)}.")
    return output


def scenarios(wbv, wbf) -> list[dict[str, Any]]:
    v = wbv["08_DECISIONES_ESCENARIOS"]
    f = wbf["08_DECISIONES_ESCENARIOS"]
    expected = (
        "Escenario", "Volumen", "Precio", "Costo directo", "Gastos operativos",
        "Ventas año 1", "Margen directo", "Margen directo %",
        "Resultado operativo simplificado", "Meses con sobrecarga", "Pico capacidad"
    )
    headers = [as_text(f.cell(36, col).value) for col in range(1, 12)]
    if tuple(headers) != expected:
        die("08_DECISIONES_ESCENARIOS fila 36: encabezados de escenarios cambiaron.")
    output = []
    for row in range(37, 41):
        name = checked_value(v, f, row, 1)
        if name in (None, ""):
            die(f"08_DECISIONES_ESCENARIOS fila {row}: escenario vacío.")
        output.append({
            "name": name,
            "volumeFactor": checked_value(v, f, row, 2),
            "priceFactor": checked_value(v, f, row, 3),
            "directCostFactor": checked_value(v, f, row, 4),
            "operatingExpenseFactor": checked_value(v, f, row, 5),
            "year1Sales": checked_value(v, f, row, 6),
            "directMargin": checked_value(v, f, row, 7),
            "directMarginPct": checked_value(v, f, row, 8),
            "simplifiedOperatingResult": checked_value(v, f, row, 9),
            "overloadMonths": checked_value(v, f, row, 10),
            "peakCapacity": checked_value(v, f, row, 11),
            "status": "ESTIMADO",
            "confidence": "Media",
            "source": "08_DECISIONES_ESCENARIOS",
        })
    return output


def decisions(wbv, wbf) -> list[dict[str, Any]]:
    v = wbv["08_DECISIONES_ESCENARIOS"]
    f = wbf["08_DECISIONES_ESCENARIOS"]
    expected = (
        "Decisión", "Mes configurado", "Mes recomendado", "Diferencia", "Estado",
        "Condición principal", "Impacto", "Acción sugerida"
    )
    headers = [as_text(f.cell(5, col).value) for col in range(1, 9)]
    if tuple(headers) != expected:
        die("08_DECISIONES_ESCENARIOS fila 5: encabezados de decisiones cambiaron.")
    output = []
    for row in range(6, 11):
        output.append({
            "name": checked_value(v, f, row, 1),
            "configuredMonth": checked_value(v, f, row, 2),
            "recommendedMonth": checked_value(v, f, row, 3),
            "differenceMonths": checked_value(v, f, row, 4),
            "decisionState": checked_value(v, f, row, 5),
            "condition": checked_value(v, f, row, 6),
            "impact": checked_value(v, f, row, 7),
            "suggestedAction": checked_value(v, f, row, 8),
            "status": "INFERIDO",
            "confidence": "Media",
            "source": "08_DECISIONES_ESCENARIOS",
        })
    return output


def pending_findings(wbv, wbf) -> list[dict[str, Any]]:
    v = wbv["06_AUDITORIA"]
    f = wbf["06_AUDITORIA"]
    expected = (
        "Prioridad", "Hallazgo", "Estado", "Impacto", "Decisión recomendada",
        "Responsable", "Fecha", "Fuente", "Riesgo", "Observación"
    )
    headers = [as_text(f.cell(23, col).value) for col in range(1, 11)]
    if tuple(headers) != expected:
        die("06_AUDITORIA fila 23: encabezados de hallazgos cambiaron.")
    output = []
    for row in range(24, 30):
        model_status = checked_value(v, f, row, 3)
        status = normalize_quality(model_status)
        output.append({
            "priority": checked_value(v, f, row, 1),
            "finding": checked_value(v, f, row, 2),
            "status": status,
            "modelStatus": model_status,
            "impact": checked_value(v, f, row, 4),
            "recommendedDecision": checked_value(v, f, row, 5),
            "owner": checked_value(v, f, row, 6),
            "due": checked_value(v, f, row, 7),
            "source": checked_value(v, f, row, 8) or "06_AUDITORIA",
            "risk": checked_value(v, f, row, 9),
            "note": checked_value(v, f, row, 10),
            "confidence": confidence_for(status),
        })
    return output


def audit_summary(wbv, wbf) -> dict[str, Any]:
    v = wbv["06_AUDITORIA"]
    f = wbf["06_AUDITORIA"]
    controls = []
    for row in range(6, 11):
        controls.append(as_text(checked_value(v, f, row, 5)))
    behaviors = []
    for row in range(15, 20):
        behaviors.append(as_text(checked_value(v, f, row, 9)))
    return {
        "integrityChecks": len(controls),
        "integrityPassed": sum(1 for value in controls if value == "PASA"),
        "behaviorTests": len(behaviors),
        "behaviorPassed": sum(1 for value in behaviors if value == "PASA"),
    }


def export_snapshot(workbook: Path, output: Path) -> dict[str, Any]:
    openpyxl, from_excel = load_openpyxl()
    wbv = openpyxl.load_workbook(workbook, read_only=False, data_only=True)
    wbf = openpyxl.load_workbook(workbook, read_only=False, data_only=False)
    validate_profile(wbf)

    product_costs, products = product_master(wbv, wbf)
    plan_sales = (
        plan_block(wbv, wbf, 5, products, from_excel)
        + plan_block(wbv, wbf, 35, products, from_excel)
    )
    flow = cash_flow(wbv, wbf, from_excel)
    scenario_rows = scenarios(wbv, wbf)
    assumption_rows = assumptions(wbv, wbf)
    decision_rows = decisions(wbv, wbf)
    pending_rows = pending_findings(wbv, wbf)
    audit = audit_summary(wbv, wbf)

    snapshot = {
        "schemaVersion": SCHEMA_VERSION,
        "meta": {
            "modelName": "MFO_EL_ERRANTE_v3_3_Decisiones_y_Escenarios",
            "modelDate": "",
            "exportedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "status": "ESTIMADO",
            "confidence": "Media",
            "source": workbook.name,
            "workbookProfile": WORKBOOK_PROFILE,
            "pendingCount": len(pending_rows),
            "decisionCount": len(decision_rows),
            **audit,
        },
        "planSales": plan_sales,
        "productCosts": product_costs,
        "cashFlow": flow,
        "scenarios": scenario_rows,
        "assumptions": assumption_rows,
        "decisions": decision_rows,
        "pending": pending_rows,
    }

    # Controles de reconciliación contra totales visibles del workbook.
    y1 = plan_sales[: 14 * 12]
    y2 = plan_sales[14 * 12 :]
    expected = {
        "y1Units": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 21, 15),
        "y1Sales": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 22, 16),
        "y1Cogs": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 23, 16),
        "y2Units": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 51, 15),
        "y2Sales": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 52, 16),
        "y2Cogs": checked_value(wbv["01_PLAN_VENTAS"], wbf["01_PLAN_VENTAS"], 53, 16),
        "endingCash": checked_value(wbv["03_RESULTADOS_CAJA"], wbf["03_RESULTADOS_CAJA"], 35, 25),
    }
    calculated = {
        "y1Units": sum(row["quantity"] for row in y1),
        "y1Sales": sum(row["sales"] for row in y1),
        "y1Cogs": sum(row["cogs"] for row in y1),
        "y2Units": sum(row["quantity"] for row in y2),
        "y2Sales": sum(row["sales"] for row in y2),
        "y2Cogs": sum(row["cogs"] for row in y2),
        "endingCash": flow[-1]["endingCash"],
    }
    mismatches = [
        key for key in expected
        if expected[key] is not None and abs(float(expected[key]) - float(calculated[key])) > 0.01
    ]
    if mismatches:
        die("El snapshot no reconcilia contra el workbook en: " + ", ".join(mismatches))
    snapshot["meta"]["reconciliation"] = "PASS"

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    return snapshot


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Exportador privado MFO v3.3 → snapshot V3.0")
    parser.add_argument("workbook", type=Path, help="Ruta local al XLSX MFO v3.3")
    parser.add_argument("--output", type=Path, default=Path("private-data/mfo_snapshot_v30.json"))
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Valida el perfil y escribe solo nombres de hojas y anclas, nunca cifras.",
    )
    parser.add_argument(
        "--inspect-output",
        type=Path,
        default=Path("private-data/mfo_profile_v33.json"),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.is_file():
        die(f"No existe el workbook: {args.workbook}")
    openpyxl, _ = load_openpyxl()
    if args.inspect:
        wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=False)
        validate_profile(wb)
        args.inspect_output.parent.mkdir(parents=True, exist_ok=True)
        args.inspect_output.write_text(
            json.dumps(inspect_profile(wb), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Perfil MFO v3.3 validado: {args.inspect_output}")
        return 0

    snapshot = export_snapshot(args.workbook, args.output)
    print(
        "Snapshot V3.0 generado y reconciliado: "
        f"{len(snapshot['planSales'])} líneas plan, "
        f"{len(snapshot['productCosts'])} SKU, "
        f"{len(snapshot['cashFlow'])} meses, "
        f"{len(snapshot['scenarios'])} escenarios, "
        f"{len(snapshot['assumptions'])} supuestos, "
        f"{len(snapshot['decisions'])} decisiones y "
        f"{len(snapshot['pending'])} pendientes."
    )
    print(f"Salida privada: {args.output}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        sys.exit(130)
